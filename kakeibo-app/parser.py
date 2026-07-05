"""
カード利用明細エクセルのパーサー。

現状はアメックス系(例: マリオット・ボンヴォイ アメリカン・エキスプレス・プレミアム・カード)の
「ご利用履歴」シート形式に対応している。

シート内のヘッダー行(例: ご利用日 / データ処理日 / ご利用内容 / カード会員様名 / 会員番号 # / 金額 ...)を
自動検出してから、その下の明細行だけを抽出する。ヘッダー行の位置は月によって多少ずれることがあるため、
「ご利用日」というセルを含む行を探して基準にする。
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import datetime

import openpyxl
import pandas as pd

# 明細シートの候補名(会社によって表記ゆれがあるため複数許容)
TRANSACTION_SHEET_CANDIDATES = ["ご利用履歴", "カードご利用明細", "利用明細"]

# ヘッダー行を特定するためのキーセル
HEADER_ANCHOR = "ご利用日"

# 必須で使う列名(これらが見つからない場合はパース失敗とする)
REQUIRED_COLUMNS = ["ご利用日", "ご利用内容", "金額"]


@dataclass
class ParsedTransaction:
    date: str  # YYYY-MM-DD
    description: str
    amount: int
    raw_hash: str  # 重複検出用のハッシュ


class ParseError(Exception):
    pass


def _find_transaction_sheet(wb: openpyxl.Workbook):
    for name in TRANSACTION_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    # 候補名に一致しなければ、HEADER_ANCHORを含むシートを探す
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True):
            if row and row[0] == HEADER_ANCHOR:
                return ws
    raise ParseError(
        "明細シートが見つかりませんでした。対応シート名: " + ", ".join(TRANSACTION_SHEET_CANDIDATES)
    )


def _find_header_row(ws) -> int:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True), start=1):
        if row and row[0] == HEADER_ANCHOR:
            return i
    raise ParseError(f"ヘッダー行(先頭セル='{HEADER_ANCHOR}')が見つかりませんでした。")


def _to_amount(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(round(value))
    s = str(value).strip()
    if s == "" or s == "None":
        return None
    s = s.replace(",", "").replace("円", "")
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def _to_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if s == "":
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def make_hash(date: str, description: str, amount: int) -> str:
    key = f"{date}|{description.strip()}|{amount}"
    return hashlib.sha256(key.encode("utf-8")).hexdigest()


def parse_excel(file) -> list[ParsedTransaction]:
    """
    アップロードされたExcelファイル(ファイルパス or file-like object)を読み込み、
    明細行のリストを返す。
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = _find_transaction_sheet(wb)
    header_row_idx = _find_header_row(ws)

    headers = [c.value for c in ws[header_row_idx]]
    # 列名インデックスを作成(空白除去)
    col_index = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        col_index[str(h).strip()] = idx

    missing = [c for c in REQUIRED_COLUMNS if c not in col_index]
    if missing:
        raise ParseError(f"必要な列が見つかりません: {missing}")

    date_col = col_index["ご利用日"]
    desc_col = col_index["ご利用内容"]
    amount_col = col_index["金額"]

    results: list[ParsedTransaction] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        date_val = row[date_col] if date_col < len(row) else None
        desc_val = row[desc_col] if desc_col < len(row) else None
        amount_val = row[amount_col] if amount_col < len(row) else None

        date_str = _to_date(date_val)
        amount = _to_amount(amount_val)
        desc = str(desc_val).strip() if desc_val is not None else ""

        if not date_str or amount is None or not desc:
            # 日付・金額・利用内容のいずれかが欠けている行はスキップ(空行や集計行など)
            continue

        h = make_hash(date_str, desc, amount)
        results.append(ParsedTransaction(date=date_str, description=desc, amount=amount, raw_hash=h))

    return results


def to_dataframe(transactions: list[ParsedTransaction]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"date": t.date, "description": t.description, "amount": t.amount, "hash": t.raw_hash}
            for t in transactions
        ]
    )
